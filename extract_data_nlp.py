import os
import pandas as pd 
import numpy as np
import spacy
import nltk
from fuzzywuzzy import fuzz
from fuzzywuzzy import process
from fuzzysearch import find_near_matches
from statistics import mean
from collections import OrderedDict
from itertools import chain
from collections import Counter
from openpyxl import load_workbook

spacy.prefer_gpu()
# NLP = spacy.load('en_core_web_sm') 
#* load our custom trained NER model
NLP = spacy.load('en_first_1.0') 


#! this is just a config file that tells what headings we want to look under. This config could come from json, dictionary, database, etc.
#* each cell contains the following: 1) region title, 2)NER label descriptions, 3)flag for whether we wish to extract NERs from that region
REGIONS_CONFIG = [['Applicant Details', ['NAME', 'COMPANY', 'ADDRESS'], 1], 
				  ['Agent Details', ['NAME', 'COMPANY', 'ADDRESS'], 1], 
				  ['Site Area', [], 0],
				  ['Materials', ['MATERIAL'], 1], 
				  ['Pedestrian and Vehicle Access, Roads and Rights of Way', [], 0]
	]

#* placeholder region list
REGIONS_LIST = []

#! this is the labels we have used to train our custome NER model
NER_LABELS = {'NAME':'PER', 'ADDRESS':'ADD', 'COMPANY':'CMP', 'MATERIAL':'MAT'}


class define_regions:
	"""this is the region class
	"""	
	def __init__(self, ip):
		self.id = ip[0]
		self.start = 0
		self.end = 0
		self.labels = ip[1]
		self.text = []
		self.extract_flag = True if ip[2] == 1 else False 
		self.name_flag = 'NAME' if 'NAME' in ip[1] else '' 
		self.names = ''
		self.address_flag = 'ADDRESS' if 'ADDRESS' in ip[1] else '' 
		self.address = ''
		self.material_flag = 'MATERIAL' if 'MATERIAL' in ip[1] else '' 
		self.materials = ''
		self.company_flag = 'COMPANY' if 'COMPANY' in ip[1] else '' 
		self.companies = ''
	
	def set_region_start(self, start):
		"""method toset the region start index inside the dataframe

		Args:
			start ([int]): [dataframe index where this region starts]
		"""		
		#* We chack is the new start index is lesser than previously defined because like in case of applicant details, regions can span multiple pages
		self.start = start+1 if start< self.start or self.start==0 else self.start

	def set_region_end(self, end):
		"""method to set the region end index inside the dataframe

		Args:
			end ([int]): [dataframe index where this region ends]
		"""		
		self.end = end
		
	def set_region_text(self, text_list):
		"""method to set the region text

		Args:
			text_list ([list]): [list containing the strings belonging to the region ]
		"""		
		self.text = text_list

	def extract_ner(self, label):
		"""method to extract the NERs for each region based on the entities present for that region inside config

		Args:
			label ([string]): [label consists of the NER label description as mentioned inside config]

		Returns:
			[string]: [string containing the extracted NERs for that region based on corresponding label]
		"""		
		#* we translate the NER label descriptions to their tags here using our lookup NER_LABELS dictionary
		ner_label= NER_LABELS.get(label)
		return_string = ''
		for i, text in enumerate(self.text):
			doc = NLP(text)
			for ent in doc.ents: 
					# print(ent.text, ent.start_char, ent.end_char, ent.label_)
					#* we append only those entities which have requested using the NER label input
				return_string = return_string + ', ' + ent.text if ent.label_ == ner_label else return_string+''
		return return_string.strip(',')
	

def show_region_details():
	"""this method is used to extract the NERs for each tag inside each of our defined regions

	#! PLEASE NOTE: since we have trainsed a custom NER model, the training data for this exercise was limited.
	#! due to which the model is overfitted for this document. as we get more sample documents, we can add more examples 
	#! or use spacy's rule based approach to generate more synthetic data for augmentine our current dataset and remodel over it
	"""	
	for region in REGIONS_LIST:
		if region.extract_flag:
			print(region.id)
			if region.name_flag != '':
				region.names = region.extract_ner(region.name_flag)
				print('names: '+ region.names)
			if region.address_flag != '':
				region.address = region.extract_ner(region.address_flag)
				print('address: '+ region.address)
			if region.material_flag != '':
				region.materials = region.extract_ner(region.material_flag)
				print('materials: '+ region.materials)
			if region.company_flag != '':
				region.companies = region.extract_ner(region.company_flag)
				print('cpmpanies: '+ region.companies)
			print('\n')


def load_config(idf):
	"""this method REGION_CONFIG and creates objects of the define_region class.
	we then set the region start and end indices by iterating over the input dataframe and getting the string scores for each region

	Args:
		idf (pandas dataframe): this is the input grouped df
	"""	
	for i, item in enumerate(REGIONS_CONFIG):
		REGIONS_LIST.append(define_regions(item))
		# print(item[0])
		for j, row in df.iterrows():
			score = getStringScore(item[0].lower(), row.text.lower())
			if score>95:
				REGIONS_LIST[i].set_region_start(j)
	
	for i, region in enumerate(REGIONS_LIST[0:len(REGIONS_LIST)-1]):
		region.set_region_end(min([item.start for item in REGIONS_LIST[i:] if item.start>region.start]))
		region.set_region_text(idf.loc[region.start:region.end]['text'].tolist())


def getStringScore(attribute, searchString):
	"""this method is used for doing string comparision using fuzzy logic 
	the string scores are calculates based on different implementations of levenshtein distance

	Args:
		attribute string: this is the input source string
		searchString string: this is the target string to compare the source string against

	Returns:
		float: this is the mean string scores for the three calculation metrics
	"""	
	stringScore = []
	# useFuzzywuzzy to get string distances and use the maximum matching value
	stringScore.append(fuzz.partial_ratio(attribute.lower(), searchString.lower()))
	stringScore.append(fuzz.token_sort_ratio(attribute.lower(), searchString.lower()))
	stringScore.append(fuzz.token_set_ratio(attribute.lower(), searchString.lower()))
	return mean(stringScore)
	

def group_regions(idf):
	"""this method is our ROI generation algorithm to have a more targeted search 
	and account for separate ROI generation rules for different types of documents

	Args:
		idf (pandas dataframe): this is the input d coming from our tesseract OCR pipeline module

	Returns:
		pandas dataframe: this dataframe contains page number, block number and text columns
	"""	
	#* we form our target df placeholder here
	newdf = pd.DataFrame(columns=['page', 'block_num', 'text'])
	block_text = ''
	block_num=0

	#* iterate over the input dataframe
	for i, row in idf.iterrows():
		if i < len(idf.index)-1:
			if block_text == '':
				block_text = row.text

			#* if the 2 text cells lie on the same y-axis level with some threshold, we merge those cells
			if row.top - (row.height*0.75) < idf.loc[i+1].top < row.top + (row.height*0.75) : #row.left + row.width + (row.height*2) > idf.loc[i+1].left and
				block_text = block_text + ' ' + idf.loc[i+1].text
				continue
			else:
				# print(block_text)
				block_num+=1
				newdf = newdf.append({'page':row.page_num, 'block_num':block_num,  'text':block_text}, ignore_index=True)
				block_text = ''
				
	return newdf
	

def extract_rule_based(idf):
	"""this is the rule based approach to extracting data from the document using a fixed templaate for each new type of document.
	if the structure and overall layout of the document is consitent, it is advisable to use rule based extraction

	Args:
		idf ([pandas dataframe]): [this is the input grouped df]
	"""	

	#* we load the template here
	wb = load_workbook('.\\template.xlsx')
	ws = wb.active
	
	model_max_row = ws.max_row
	for k in range(2, model_max_row + 1):
		try:
			# field_name is the region we wish to extract from
			field_name = ws.cell(row = k, column=1).value
			# parent is the label we wish to display against the data
			parent = ws.cell(row=k, column=2).value
			# child is the labels which lie in the field, eg. address line 1, address line 2, country etc
			child = ws.cell(row = k, column = 3).value
			child = child.split(',') if child is not None else []
			data_list = []

			for i, region in enumerate(REGIONS_LIST):
				#* iterate over the regions to find the one corresponding to what has been modelled in the template
				if getStringScore(field_name.lower(), region.id) > 90:
					#* reset the child index for that region. child index is used so that we dont keep matching the same child over and over again
					ch_ind = 0
					for ch in child:
						for text_ind, text_row in enumerate(region.text):
							#* this is done in order to exclude the children we have already extracted since we use fuzzy searching below
							if text_ind > ch_ind:
								indices = find_near_matches(ch.lower(), text_row.lower(), max_insertions=1, max_deletions=1, max_substitutions=1)
								if len(indices) > 0:
									ch_ind = text_ind
									label_end = indices[0].end
									data_list.append(text_row[label_end:])
									break
			print(f'{field_name} - {parent} -->'+ ''.join(data_list))
		except Exception as e:
			print(e)
			pass
	wb.close()


if __name__ == '__main__':
	"""here we load the op excel file into a dataframme from the previous module
	merge line region cells to obtain our regions of interest.  
	"""	
	df = pd.read_excel('.\\op.xlsx', engine='openpyxl')

	#* merge text cells to form line level ROI. different ROI rules can be set up here for different images and templates
	df = group_regions(df)
	df.to_excel('clean_op.xlsx')

	#* this is the config file that tells the application what headings to split the data into.
	#* doing this helps us narrow our results better
	load_config(idf=df)

	#* method to display the information using NLP
	show_region_details()
	#* method to display the information using a template/rule based approach
	extract_rule_based(df)